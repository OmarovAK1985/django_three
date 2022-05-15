from books.models import Author, Book
from django.core.management.base import BaseCommand
import openpyxl
import os


class Command(BaseCommand):
    def add_arguments(self, parser):
        pass

    def handle(self, *args, **options):
        file = os.path.join(os.getcwd(), 'books.xlsx')
        excel_reader = openpyxl.load_workbook(file)
        list_excel = excel_reader['list']
        list_authors = []
        for i in range(2, list_excel.max_row + 1):
            count = list_excel.cell(column=2, row=i).value
            author = list_excel.cell(column=2, row=i).value
            if count not in list_authors:
                list_authors.append(author)

        list_authors_in_model = []
        if len(Author.objects.all()) > 0:
            for i in Author.objects.all():
                list_authors_in_model.append(str(i))

        if len(Author.objects.all()) == 0:
            for i in list_authors:
                Author.objects.create(author=i)
        elif len(Author.objects.all()) > 0:
            for i in list_authors:
                if i not in list_authors_in_model:
                    Author.objects.create(author=i)
        my_list = []
        for i in range(2, list_excel.max_row + 1):
            name_book = list_excel.cell(column=1, row=i).value
            author = list_excel.cell(column=2, row=i).value
            date_publication = list_excel.cell(column=3, row=i).value
            my_dict = dict()
            my_dict.setdefault('name_book', str(name_book))
            my_dict.setdefault('author', author)
            my_dict.setdefault('date_publication', date_publication)
            my_list.append(my_dict)

        list_book_in_model = []
        for i in my_list:
            for key, val in i.items():
                if len(Book.objects.all()) == 0:
                    Book.objects.create(name_book=i['name_book'], date_publication=i['date_publication'], author=Author.objects.get(author=i['author']))
                elif len(Book.objects.all()) > 0:
                    for k in Book.objects.all():
                        list_book_in_model.append(str(k))
                    if i['name_book'] not in list_book_in_model:
                        Book.objects.create(name_book=i['name_book'], date_publication=i['date_publication'],
                                            author=Author.objects.get(author=i['author']))









